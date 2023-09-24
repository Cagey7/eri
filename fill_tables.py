from datetime import datetime
import os
import calendar
from openpyxl import Workbook
from new_data_automation import Automation

class FillTable(Automation):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.excel_path = os.path.join(os.getcwd())


    def consumer_price_index_fill_table(self, start_month=None, start_year=None, end_month=None, end_year=None):
        excel_file_names = [["consumer_price_index1.xlsx", "отчетный период к соответствующему периоду прошлого года"], 
                            ["consumer_price_index2.xlsx", "отчетный период к предыдущему периоду"]]
        
        for name in excel_file_names:
            workbook = Workbook()
            sheet = workbook.active

            excel_path = os.path.join(self.excel_path, name[0])

            if end_month is None or end_year is None:
                current_date = datetime.now()
                end_month = current_date.month
                end_year = current_date.year
            
            if start_month is None or start_year is None:
                current_date = datetime.now()
                start_month = 12
                start_year = current_date.year - 2
            
            start_last_day = calendar.monthrange(start_year, start_month)[1]
            end_last_day = calendar.monthrange(end_year, end_month)[1]

            months = [
                ["янв", 'Январь'], 
                ["фев", 'Февраль'], 
                ["мар", 'Март'], 
                ["апр", 'Апрель'], 
                ["май", 'Май'], 
                ["июн", 'Июнь'], 
                ["июл", 'Июль'], 
                ["авг", 'Август'], 
                ["сен", 'Сентябрь'], 
                ["окт", 'Октябрь'], 
                ["ноя", 'Ноябрь'], 
                ["дек", 'Декабрь']
            ]

            bull_name = [
                ["Потребительские цены", 'Товары и услуги'], 
                ["Продовольственные товары", 'Продовольственные товары'], 
                ["Непродовольственные товары", 'Непродовольственные товары'], 
                ["Платные услуги", 'Платные услуги'], 
                ["Цены производителей", 'Промышленность']
            ]


            # Создание мест списку для поиска
            condition = ",".join(["%s"] * len(bull_name))


            

            join_views = f"""
            SELECT 
                activity_type, 
                value, 
                description, 
                created_at
            FROM 
                consumer_price_index_month 
            WHERE 
                region = 'РЕСПУБЛИКА КАЗАХСТАН' 
                AND periods_correlation = '{name[1]}'  
                AND activity_type IN ({condition}) 
                AND created_at BETWEEN '{start_year}-{start_month}-{start_last_day}' AND '{end_year}-{end_month}-{end_last_day}'
            UNION
            SELECT 
                activity_type, 
                value, 
                description, 
                created_at
            FROM 
                producer_price_index_month 
            WHERE 
                region = 'РЕСПУБЛИКА КАЗАХСТАН' 
                AND periods_correlation = '{name[1]}' 
                AND countries = 'Всего'
                AND activity_type = 'Промышленность'
                AND industrial_products_list = 'Всего'
                AND created_at BETWEEN '{start_year}-{start_month}-{start_last_day}' AND '{end_year}-{end_month}-{end_last_day}'
            ORDER BY 
                created_at;
            """

            self.cur.execute(join_views, [item[1] for item in bull_name])
            sorted_data = self.cur.fetchall()
            dates = [item[3] for item in sorted_data]
            latest_month = max(dates).month
            latest_year = max(dates).year

            if latest_month < end_month:
                end_month = latest_month
            if latest_year < end_year:
                end_year = latest_year

            
            months_list = [""]

            data_for_excel = []
            for bull_name_i, economic_index in enumerate(bull_name):
                index_data = []
                index_data.append(economic_index[0])
                for year in range(start_year, end_year+1):
                    for i, m in enumerate(months, start=1):
                        if year == start_year and i < start_month: continue
                        if year == end_year and i > end_month: continue
                        # получешние месяцев для заголовка
                        if bull_name_i == 0:
                            months_list.append(m[0])
                        
                        data_not_exist = True
                        for data in sorted_data:
                            t_month, t_year, _ = data[2].split()
                            t_year = int(t_year)
                            if economic_index[1] == data[0] and t_month == m[1] and t_year == year:
                                index_data.append(float(data[1]))
                                data_not_exist = False
                        if data_not_exist:
                            index_data.append("")
                data_for_excel.append(index_data)

            # заголовок
            year_list = list(range(start_year, end_year + 1))
            start_cell = 2
            end_cell = 12-start_month+start_cell
            sheet.cell(row=1, column=1, value="Показатели")
            for i, year in enumerate(year_list):
                sheet.cell(row=1, column=start_cell, value=year)
                if start_year == end_year and start_month == end_month: break
                sheet.merge_cells(start_row=1, start_column=start_cell, end_row=1, end_column=end_cell)
                start_cell = end_cell+1
                if i+2 == len(year_list):
                    end_cell += end_month
                else:
                    end_cell += 12



            sheet.append(months_list)
            for data in data_for_excel:
                sheet.append(data)
            workbook.save(excel_path)
            workbook.close()


fillTable = FillTable("testtest", "postgres", "123456")

fillTable.consumer_price_index_fill_table()
fillTable.db_disconnect()